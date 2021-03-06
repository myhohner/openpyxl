import openpyxl,os,re,asyncio,sys,io
from search import Search
from datetime import datetime
import multiprocessing as mp
#sys.stdout=io.TextIOWrapper(sys.stdout.buffer,encoding='gb18030') 
class Action:


    def pool_action(self,b,url):
        return b.search_test(url)


    def multicore(self,url_list):
        self.pool = mp.Pool()
        self.b=Search()
        result_dict=self.pool.map(self.b.search_test,url_list)#获得html列表
        result_dict=dict(result_dict)
        return result_dict

    def check_crawl(self,result_dict,data_list):
        dict_update=self.b.check_crawl(result_dict,data_list)
        return dict_update

    def selenium_test(self,dict_update,data_list):
        for url,num in dict_update.items():
            if num=='2' or num=='3':
                res_html=self.pool.apply_async(self.b.selenium,(url,))
                res_html=res_html.get()
                for i in data_list:
                    if re.search(i,res_html):
                        dict_update[url]='1'
        #os.system('taskkill /im chromedriver.exe /F')
        #os.system('taskkill /im chrome.exe /F')
        self.pool.close()
        return dict_update

#注意：excel网址不能重复
if __name__=='__main__':
    time1=datetime.now()
    path=os.path.dirname(__file__)
    workbook = openpyxl.load_workbook(path+'/'+'test.xlsx')#读入文件的路径 
    booksheet=workbook.active
    max_row = booksheet.max_row
    values=booksheet.iter_rows(min_row=2,min_col=2,max_row=max_row,max_col=2)
    url_list=['%s'%i[0].value for i in values] #获取查询网址
    '''
    url_list=['http://dy.163.com/v2/article/detail/FECD42PS053469JX.html',
'http://emcreative.eastmoney.com//Fortune/V/Share_ArticleDetail/20200605203712449232300',
'http://www.toutiao.com/i6837381215624888840/',
'http://bigdata-s3.wmcloud.com/newsmbl/vaI5WfrER6IhhfiqeYqi6Q',
'http://www.sznews.com/eating/content/2020-06/05/content_23223474.htm',
'http://www.yidianzixun.com/article/0PXvWfAd']
'''

    data_list=['京基','农']

    a=Search()
    a.selenium_setup()

    '''
    #普通爬虫
    #result_list=a.search(url_list,data_list)#获得查询结果
    '''

    #asyncio
    result_list=asyncio.run(a.task_manager(url_list,data_list)) #3.7使用此方法启动协程
    print(result_list)#测试
    a.tearDown()
    '''
    #分布式爬虫
    a=Action()
    result_dict=a.multicore(url_list)
    dict_update_origin=a.check_crawl(result_dict,data_list)
    dict_update=a.selenium_test(dict_update_origin,data_list)


    #url_list=['http://www.jlmbbz.com/news/dyxw/20200612/137055.html']
    data_list=['京基','农']
    #a=Search()
    #a.selenium_setup()
    #result_list=a.search(url_list,data_list)#获得查询结果

    a=Action()
    #b.selenium_setup()
    #result_list=[pool.apply_async(b.search_test,args=(i,)) for i in url_list]#获得查询结果
    result_dict=a.multicore(url_list)
    dict_update=a.check_crawl(result_dict,data_list)
    dict_update=a.selenium_test(dict_update,data_list)
>>>>>>> bf36c3b9f85edd3ea322cd6a23bedcaa967f6f44
    result_list=[]
    for key in dict_update.keys():
        value=dict_update[key]
        result_list.append(value)#将dict值转为list
<<<<<<< HEAD
    '''

    fill_colums=3
    ws = booksheet.insert_cols(fill_colums)
    for index,row in enumerate(booksheet.rows):
        if index!=0:
            row[fill_colums-1].value=result_list[index-1]
    workbook.save(path+'/'+'test.xlsx')
    time2=datetime.now()
    print(time2-time1)
